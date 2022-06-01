"""Create excel file with network live cell data."""

import os

from openpyxl import Workbook


def create_excel(network_live_data):
    """
    Create excel file with network live cell data.

    Args:
        network_live_data: dict

    Returns:
        string
    """
    file_path = 'network_live/logs/kcell_cells.xlsx'
    os.remove(file_path)
    headers = {
        'lte': [
            'SubNetwork',
            'NODEID',
            'SITENAME',
            'EUTRANCELL',
            'TAC',
            'CELLID',
            'ECI',
            'EARFCNDL',
            'QRXLEVMIN',
            'LATITUDE',
            'LONGITUDE',
            'ADMINISTRATIVESTATE',
            'RACHROOTSEQUENCE',
            'PHYSICALCELLID',
            'IPADDRESS',
            'VENDOR',
            'INSERTDATE',
            'OSS',
        ],
        'wcdma': [
            'OPERATOR',
            'RNCID',
            'RNCNAME',
            'SITENAME',
            'UTRANCELL',
            'LOCALCELLID',
            'UARFCNDL',
            'UARFCNUL',
            'PSC',
            'LAC',
            'RAC',
            'SAC',
            'URALIST',
            'PRIMARYCPICHPOWER',
            'MAXIMUMTRANSMISSIONPOWER',
            'IUBLINK',
            'MOCNCELLPROFILE',
            'ADMINISTRATIVESTATE',
            'IPADDRESS',
            'VENDOR',
            'INSERTDATE',
            'OSS',
        ],
        'gsm': [
            'OPERATOR',
            'BSCID',
            'BSCNAME',
            'SITENAME',
            'CELL',
            'BCC',
            'NCC',
            'LAC',
            'CELLID',
            'BCCH',
            'HSN',
            'MAIO',
            'TCHFREQS',
            'CELLSTATE',
            'VENDOR',
            'INSERTDATE',
            'OSS',
        ],
        'nr': [
            'SUBNETWORK',
            'GNBID',
            'SITENAME',
            'CELLNAME',
            'CELLID',
            'CELLSTATE',
            'NCI',
            'NRPCI',
            'NRTAC',
            'RACHROOTSEQUENCE',
            'QRXLEVMIN',
            'ARFCNDL',
            'BSCHANNELBWDL',
            'CONFIGUREDMAXTXPOWER',
            'IPADDRESS',
            'VENDOR',
            'INSERTDATE',
            'OSS',
        ],
    }

    work_book = Workbook()

    for technology, cell_data in network_live_data.items():
        row = 1
        column = 1
        work_sheet = work_book.create_sheet(technology.upper())
        for header in headers[technology]:
            work_sheet.cell(row=row, column=column, value=header)
            column += 1

        row += 1
        column = 1

        for cell_values in cell_data:
            for value in cell_values:
                work_sheet.cell(row=row, column=column, value=value)
                column += 1
            column = 1
            row += 1

    work_book.remove_sheet(work_book.get_sheet_by_name('Sheet'))
    work_book.save(file_path)
    return file_path
