"""Fill excel file with all sites."""

import os

from openpyxl import Workbook


def fill_table(sql_data):
    """
    Fill excel file with all sites.

    Args:
        sql_data: list

    Returns:
        string
    """
    table_path = 'site_count/logs/site_data.xlsx'
    if os.path.exists(table_path):
        os.remove(table_path)

    work_book = Workbook()
    work_sheet = work_book.active
    work_sheet.column_dimensions['A'].width = 40
    work_sheet.column_dimensions['H'].width = 25

    headers = ('Site', 'Operator', 'GSM', 'WCDMA', 'LTE', 'NR5G', 'Vendor', 'Region')
    for header in headers:
        work_sheet.cell(row=1, column=headers.index(header) + 1, value=header)

    row_index = 2
    for row in sql_data:
        col_index = 1
        for cell_value in row:
            work_sheet.cell(row=row_index, column=col_index, value=cell_value)
            col_index += 1
        row_index += 1

    work_book.save(table_path)
    return table_path
